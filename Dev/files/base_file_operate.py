#!python3
# -*- coding: utf-8 -*-

"""
base_file_operate.py

文件的基础操作。

该模块包含以下功能：
1. 取得路径filepath下的所有文件名(不包括子目录下的文件)
2. 取得路径filepath下的所有文件名(包括子目录下的文件)
3. 取得excel的全部数据，将其保存到一个字典中

Author: niu0217
Date: 2023-12-20
"""

import os
import openpyxl


def get_filename(folder):
    """
    取得路径filepath下的所有文件名(不包括子目录下的文件)

    Args:
        folder (str): 想要查找的文件路径

    Returns:
        list: 保存了当前路径下的所有文件名字
    """

    filenames_list = []
    for filename in os.listdir(folder):
        filenames_list.append(filename)

    return filenames_list


def get_filename_recursion(filepath):
    """
    取得路径filepath下的所有文件名(包括子目录下的文件)

    Args:
        filepath (str): 想要查找的文件路径

    Returns:
        list: 保存了当前路径下的所有文件名字
    """

    filenames_list = []
    for foldername, subdfolders, filenames in os.walk(filepath):
        for filename in filenames:
            filenames_list.append(filename)

    return filenames_list


def get_excel_data(filename):
    """取得excel的全部数据，将其保存到一个字典中"""
    excel_data_dict = {}
    wb = openpyxl.load_workbook(filename)
    print(wb.sheetnames)
    for tablename in wb.sheetnames:
        sheet = wb[tablename]
        table_data_list = []
        # print(f"row: {sheet.max_row}\ncolumn: {sheet.max_column}")
        for rowindex in range(1, sheet.max_row + 1):
            for columnindex in range(1, sheet.max_column + 1):
                # print(sheet.cell(row=rowindex, column=columnindex).value, end=' ')
                table_data_list.append(sheet.cell(row=rowindex, column=columnindex).value)
            # print('')
        excel_data_dict[tablename] = table_data_list

    return excel_data_dict



if __name__ == '__main__':
    # result_list = get_filename_recursion('/Users/niu0217/niuGithub/Documents')
    # print(result_list)
    get_excel_data('./sources/example.xlsx')
